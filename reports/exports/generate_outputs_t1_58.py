#!/usr/bin/env python3
"""End-to-end deliverables export (tournaments 1-58).

Exports:
- SQL  -> data/sql/generated/
- Excel -> data/extracts/generated/

Tables (6):
- Entries
- DrawPlayers
- Matches
- PlayerSuspensions
- PointsHistory
- WeeklyRanking

Pipeline order per tournament:
    Entries -> DrawPlayers -> Matches -> PlayerSuspensions -> PointsHistory
Then globally:
    WeeklyRanking (all weeks in ranking range)
"""

from __future__ import annotations

import math
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook

# Upstream modules
from scripts.generation.generate_entries import generate_entries
from scripts.generation.generate_draw_players import generate_draw_players
from scripts.generation.generate_matches import generate_knockout_matches
from src.modules.generate_player_suspensions import generate_player_suspensions
from src.modules.calculate_points_history import calculate_points_history
from src.modules.calculate_weekly_ranking import calculate_weekly_ranking
from src.modules.rules_engine import (
    AgeCategoryRule,
    has_super_tiebreak,
    should_seed_tournament,
    RANKING_START_YEAR,
    RANKING_START_WEEK,
    RANKING_END_YEAR,
    RANKING_END_WEEK,
)
from src.modules.ranking_window import IsoWeek, add_iso_weeks


# ============================================================
# CONFIG
# ============================================================

TOURNAMENT_ID_MIN = 1
TOURNAMENT_ID_MAX = 58

REPO_ROOT = Path(__file__).resolve().parents[2]

SQL_OUT_DIR = REPO_ROOT / "data" / "sql" / "generated"
XLSX_OUT_DIR = REPO_ROOT / "data" / "extracts" / "generated"

SQL_BLANK_LINES_BETWEEN_TABLE_BLOCKS = 3


# ============================================================
# UTILITIES
# ============================================================

def _now_utc_naive() -> datetime:
    """Timezone-naive UTC now (Excel-compatible)."""
    return datetime.now(timezone.utc).replace(tzinfo=None)


def _ensure_dirs() -> None:
    SQL_OUT_DIR.mkdir(parents=True, exist_ok=True)
    XLSX_OUT_DIR.mkdir(parents=True, exist_ok=True)


def _strip_tz(v: Any) -> Any:
    """Strip tzinfo from datetime objects for Excel compatibility."""
    if isinstance(v, datetime) and v.tzinfo is not None:
        return v.replace(tzinfo=None)
    return v


def _to_short_date(v: Any) -> Any:
    """Convert datetime to date (short date) for DATE schema columns."""
    if isinstance(v, datetime):
        return v.date()
    return v


def _normalise_rows(
    rows: List[Dict[str, Any]],
) -> Tuple[List[str], List[Dict[str, Any]]]:
    if not rows:
        return [], []
    cols = list(rows[0].keys())
    seen = set(cols)
    for r in rows[1:]:
        for k in r.keys():
            if k not in seen:
                cols.append(k)
                seen.add(k)
    return cols, rows


def _sql_literal(v: Any) -> str:
    if v is None:
        return "NULL"
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, (int, float)):
        return str(v)
    if isinstance(v, datetime):
        return f"'{v.strftime('%Y-%m-%d %H:%M:%S')}'"
    if isinstance(v, date):
        return f"'{v.strftime('%Y-%m-%d')}'"
    s = str(v).replace("'", "''")
    return f"'{s}'"


def _xlsx_rows_by_header(path: Path) -> List[Dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    out: List[Dict[str, Any]] = []
    for r in rows[1:]:
        d: Dict[str, Any] = {}
        for i, k in enumerate(header):
            if not k:
                continue
            d[k] = r[i] if i < len(r) else None
        out.append(d)
    return out


def _next_power_of_2(n: int) -> int:
    """Return the smallest power of 2 >= n."""
    if n <= 0:
        return 1
    return 1 << (n - 1).bit_length()


# ============================================================
# DATA LOADERS
# ============================================================

def _load_players() -> List[Dict[str, Any]]:
    return _xlsx_rows_by_header(
        REPO_ROOT / "data" / "extracts" / "prerequisites" / "Players.xlsx"
    )


def _load_draws() -> List[Dict[str, Any]]:
    return _xlsx_rows_by_header(
        REPO_ROOT / "data" / "extracts" / "prerequisites" / "Draws.xlsx"
    )


def _load_tournaments() -> List[Dict[str, Any]]:
    return _xlsx_rows_by_header(
        REPO_ROOT / "data" / "extracts" / "prerequisites" / "Tournaments.xlsx"
    )


def _load_age_category_rules() -> Tuple[AgeCategoryRule, ...]:
    rows = _xlsx_rows_by_header(
        REPO_ROOT / "data" / "extracts" / "prerequisites" / "AgeCategory.xlsx"
    )
    return tuple(
        AgeCategoryRule(
            age_category_id=int(r["age_category_id"]),
            min_age=int(r["min_age"]),
            max_age=int(r["max_age"]),
        )
        for r in rows
    )


def _load_points_rules() -> Dict[Tuple[int, int], int]:
    """Load PointsRules as {(category_id, stage_result_id): points}."""
    rows = _xlsx_rows_by_header(
        REPO_ROOT / "data" / "extracts" / "prerequisites" / "PointsRules.xlsx"
    )
    return {
        (int(r["category_id"]), int(r["stage_result_id"])): int(r["points"])
        for r in rows
    }


def _load_seeding_rules() -> List[Dict[str, Any]]:
    return _xlsx_rows_by_header(
        REPO_ROOT / "data" / "extracts" / "prerequisites" / "SeedingRules.xlsx"
    )


def _build_tournament_map(
    tournaments: List[Dict[str, Any]],
) -> Dict[int, Dict[str, Any]]:
    """tournament_id -> full tournament row."""
    return {int(t["tournament_id"]): t for t in tournaments}


def _build_draw_map(draws: List[Dict[str, Any]]) -> Dict[int, Dict[str, Any]]:
    """draw_id -> full draw row."""
    return {int(d["draw_id"]): d for d in draws}


def _build_draws_by_tournament(
    draws: List[Dict[str, Any]],
) -> Dict[int, List[Dict[str, Any]]]:
    """tournament_id -> list of draw rows."""
    out: Dict[int, List[Dict[str, Any]]] = {}
    for d in draws:
        tid = int(d["tournament_id"])
        out.setdefault(tid, []).append(d)
    return out


def _build_player_gender_map(
    players: List[Dict[str, Any]],
) -> Dict[int, int]:
    """player_id -> gender_id."""
    return {int(p["player_id"]): int(p["gender_id"]) for p in players}


def _seeding_rule_for_num_players(
    num_players: int,
    seeding_rules_rows: List[Dict[str, Any]],
) -> Dict[str, int]:
    """Build the seeding_rules dict expected by generate_draw_players."""
    draw_size = _next_power_of_2(num_players)
    num_seeds = 0
    for sr in seeding_rules_rows:
        if int(sr["min_players"]) <= num_players <= int(sr["max_players"]):
            num_seeds = int(sr["num_seeds"])
            break
    num_byes = draw_size - num_players
    return {
        "draw_size": draw_size,
        "num_seeds": num_seeds,
        "num_byes": num_byes,
    }


# ============================================================
# WRITE FUNCTIONS
# ============================================================

def write_sql_inserts(
    *,
    out_path: Path,
    table_name: str,
    rows: List[Dict[str, Any]],
) -> None:
    cols, rows2 = _normalise_rows(rows)
    lines: List[str] = [
        "-- Generated by reports/exports/generate_outputs_t1_58.py",
        f"-- Table: {table_name}",
        f"-- Rows: {len(rows2)}",
        "",
    ]
    if not rows2:
        lines.append(f"-- No rows generated for {table_name}")
        lines.append("")
    else:
        col_list = ", ".join(cols)
        for r in rows2:
            values = ", ".join(_sql_literal(r.get(c)) for c in cols)
            lines.append(
                f"INSERT INTO {table_name} ({col_list}) VALUES ({values});"
            )
        lines.extend([""] * SQL_BLANK_LINES_BETWEEN_TABLE_BLOCKS)
    out_path.write_text(
        "\n".join(lines).rstrip("\n") + "\n", encoding="utf-8"
    )


def write_xlsx(
    *,
    out_path: Path,
    sheet_name: str,
    rows: List[Dict[str, Any]],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    cols, rows2 = _normalise_rows(rows)
    ws.append(cols)
    for r in rows2:
        ws.append([_strip_tz(r.get(c)) for c in cols])
    wb.save(out_path)


# ============================================================
# DELIVERABLES
# ============================================================

@dataclass(frozen=True)
class Deliverable:
    table_name: str
    sql_name: str      # lowercase (matches files on disk)
    xlsx_name: str     # CamelCase (matches files on disk)


DELIVERABLES: Tuple[Deliverable, ...] = (
    Deliverable("Entries", "entries.sql", "Entries.xlsx"),
    Deliverable("DrawPlayers", "drawplayers.sql", "DrawPlayers.xlsx"),
    Deliverable("Matches", "matches.sql", "Matches.xlsx"),
    Deliverable("PlayerSuspensions", "playersuspensions.sql", "PlayerSuspensions.xlsx"),
    Deliverable("PointsHistory", "pointshistory.sql", "PointsHistory.xlsx"),
    Deliverable("WeeklyRanking", "weeklyranking.sql", "WeeklyRanking.xlsx"),
)


# ============================================================
# GENERATION PIPELINE
# ============================================================

def generate_all_rows_for_tournaments(
    tournament_ids: List[int],
) -> Dict[str, List[Dict[str, Any]]]:

    # Load all prerequisite data
    players = _load_players()
    draws_all = _load_draws()
    tournaments_all = _load_tournaments()
    age_category_rules = _load_age_category_rules()
    points_rules = _load_points_rules()
    seeding_rules_rows = _load_seeding_rules()

    tournament_map = _build_tournament_map(tournaments_all)
    draw_map = _build_draw_map(draws_all)
    draws_by_tournament = _build_draws_by_tournament(draws_all)
    player_gender = _build_player_gender_map(players)

    # Accumulator for all output rows
    rows: Dict[str, List[Dict[str, Any]]] = {
        d.table_name: [] for d in DELIVERABLES
    }

    # Counters
    next_entry_id = 1
    next_match_id = 1
    next_ph_id = 1
    next_suspension_id = 1

    # ============================================================
    # PER-TOURNAMENT PIPELINE
    # ============================================================

    for tid in tournament_ids:
        t_info = tournament_map.get(tid)
        if t_info is None:
            raise RuntimeError(f"Missing tournament_id={tid} in Tournaments.xlsx")

        t_start = t_info["start_date"]
        t_end = t_info["end_date"]
        t_year = int(t_info["tournament_year"])
        category_id = int(t_info["category_id"])

        # Ensure DATE columns are short date (not datetime)
        t_start_date = t_start.date() if isinstance(t_start, datetime) else t_start
        t_end_date = t_end.date() if isinstance(t_end, datetime) else t_end

        tournament_draws = draws_by_tournament.get(tid, [])
        if not tournament_draws:
            continue

        disable_seeding = not should_seed_tournament(tid)

        # --- STEP 1: Entries ---
        generated_entries = generate_entries(
            tournament_id=tid,
            tournament_start_date=t_start,
            eligible_players={},
            next_entry_id=next_entry_id,
            tournament_year=t_year,
            players=players,
            draws=draws_all,
            age_category_rules=age_category_rules,
            player_suspensions=rows["PlayerSuspensions"],
        )
        rows["Entries"].extend(generated_entries)
        if generated_entries:
            next_entry_id = max(
                int(e["entry_id"]) for e in generated_entries
            ) + 1

        # Group entries by (age_category_id, gender_id) -> draw_id
        entries_by_draw: Dict[int, List[Dict[str, Any]]] = {}
        for draw_row in tournament_draws:
            did = int(draw_row["draw_id"])
            ac_id = int(draw_row["age_category_id"])
            g_id = int(draw_row["gender_id"])
            draw_entries = [
                e for e in generated_entries
                if int(e["age_category_id"]) == ac_id
                and int(e["gender_id"]) == g_id
            ]
            if draw_entries:
                entries_by_draw[did] = draw_entries

        # --- STEP 2: DrawPlayers ---
        draw_players_by_draw: Dict[int, List[Dict[str, Any]]] = {}

        for draw_row in tournament_draws:
            did = int(draw_row["draw_id"])
            draw_entries = entries_by_draw.get(did, [])
            if not draw_entries:
                continue

            num_players = len(draw_entries)
            draw_generated_at = draw_row.get("draw_generated_at")
            if draw_generated_at is None:
                draw_generated_at = _now_utc_naive()

            sr = _seeding_rule_for_num_players(num_players, seeding_rules_rows)

            dp_rows = generate_draw_players(
                draw_id=did,
                entries=draw_entries,
                draw_generated_timestamp=draw_generated_at,
                seeding_rules=[sr],
                disable_seeding=disable_seeding,
            )

            rows["DrawPlayers"].extend(dp_rows)
            draw_players_by_draw[did] = dp_rows

        # --- STEP 3: Matches ---
        draw_to_tournament: Dict[int, int] = {}
        matches_by_draw: Dict[int, List[Dict[str, Any]]] = {}

        for draw_row in tournament_draws:
            did = int(draw_row["draw_id"])
            draw_to_tournament[did] = tid
            ac_id = int(draw_row["age_category_id"])
            g_id = int(draw_row["gender_id"])

            dps = draw_players_by_draw.get(did, [])
            if not dps:
                continue

            use_supertiebreak = bool(
                draw_row.get("has_supertiebreak",
                             has_super_tiebreak(ac_id, g_id))
            )

            generated_matches = generate_knockout_matches(
                draw_id=did,
                draw_players=dps,
                tournament_start_date=t_start,
                has_supertiebreak=use_supertiebreak,
                next_match_id=next_match_id,
            )

            # Ensure match_date is short date (DATE, not TIMESTAMP)
            for m in generated_matches:
                if "match_date" in m:
                    m["match_date"] = _to_short_date(m["match_date"])

            rows["Matches"].extend(generated_matches)
            matches_by_draw[did] = generated_matches
            if generated_matches:
                next_match_id = max(
                    int(m["match_id"]) for m in generated_matches
                ) + 1

        # --- STEP 4: PlayerSuspensions ---
        generated_susp = generate_player_suspensions(
            matches=rows["Matches"],
            draw_to_tournament_id=draw_to_tournament,
            starting_suspension_id=next_suspension_id,
            created_at=_now_utc_naive(),
        )

        new_susp_rows = [
            {
                "suspension_id": r.suspension_id,
                "player_id": r.player_id,
                "tournament_id": r.tournament_id,
                "reason_match_status_id": r.reason_match_status_id,
                "suspension_start": r.suspension_start,
                "suspension_end": r.suspension_end,
            }
            for r in generated_susp
        ]

        # Replace entire list (dedup handled inside generate_player_suspensions)
        rows["PlayerSuspensions"] = new_susp_rows
        if new_susp_rows:
            next_suspension_id = max(
                int(s["suspension_id"]) for s in new_susp_rows
            ) + 1

        # --- STEP 5: PointsHistory ---
        for draw_row in tournament_draws:
            did = int(draw_row["draw_id"])
            ac_id = int(draw_row["age_category_id"])

            draw_matches = matches_by_draw.get(did, [])
            dps = draw_players_by_draw.get(did, [])
            if not draw_matches or not dps:
                continue

            player_ids_in_draw = [int(dp["player_id"]) for dp in dps]

            # Ensure tournament_end_date is short date
            ph_rows = calculate_points_history(
                draw_id=did,
                tournament_id=tid,
                age_category_id=ac_id,
                category_id=category_id,
                points_rules=points_rules,
                tournament_end_date=t_end_date,
                matches=draw_matches,
                player_ids=player_ids_in_draw,
                next_ph_id=next_ph_id,
            )

            rows["PointsHistory"].extend(ph_rows)
            if ph_rows:
                next_ph_id = max(int(p["id"]) for p in ph_rows) + 1

    # ============================================================
    # WEEKLY RANKING (all weeks in configured range)
    # ============================================================

    # Build tournament metadata for weekly ranking: {tid: {year, week}}
    tournament_week_meta: Dict[int, Dict[str, int]] = {}
    for tid_key in tournament_ids:
        t_info = tournament_map.get(tid_key)
        if t_info is None:
            continue
        tournament_week_meta[tid_key] = {
            "year": int(t_info["tournament_year"]),
            "week": int(t_info["tournament_week"]),
        }

    # Generate rankings for each week in range
    current_week = IsoWeek(RANKING_START_YEAR, RANKING_START_WEEK)
    end_week = IsoWeek(RANKING_END_YEAR, RANKING_END_WEEK)

    while (
        current_week.iso_year < end_week.iso_year
        or (
            current_week.iso_year == end_week.iso_year
            and current_week.iso_week <= end_week.iso_week
        )
    ):
        wr_rows = calculate_weekly_ranking(
            ranking_year=current_week.iso_year,
            ranking_week=current_week.iso_week,
            points_history=rows["PointsHistory"],
            tournaments=tournament_week_meta,
            player_gender=player_gender,
        )
        rows["WeeklyRanking"].extend(wr_rows)

        current_week = add_iso_weeks(
            current_week.iso_year, current_week.iso_week, 1
        )

    return rows


# ============================================================
# MAIN
# ============================================================

def main() -> None:
    _ensure_dirs()

    tournament_ids = list(range(TOURNAMENT_ID_MIN, TOURNAMENT_ID_MAX + 1))
    rows_by_table = generate_all_rows_for_tournaments(tournament_ids)

    for d in DELIVERABLES:
        table_rows = rows_by_table[d.table_name]

        write_sql_inserts(
            out_path=SQL_OUT_DIR / d.sql_name,
            table_name=d.table_name,
            rows=table_rows,
        )

        write_xlsx(
            out_path=XLSX_OUT_DIR / d.xlsx_name,
            sheet_name=d.table_name,
            rows=table_rows,
        )

    print(f"Export complete: {TOURNAMENT_ID_MIN}-{TOURNAMENT_ID_MAX}")
    for d in DELIVERABLES:
        count = len(rows_by_table[d.table_name])
        print(f"  {d.table_name}: {count} rows")

    (SQL_OUT_DIR / "_export_meta.txt").write_text(
        "\n".join([
            f"generated_at_utc={_now_utc_naive().strftime('%Y-%m-%d %H:%M:%S')}",
            f"tournament_id_min={TOURNAMENT_ID_MIN}",
            f"tournament_id_max={TOURNAMENT_ID_MAX}",
        ]) + "\n",
        encoding="utf-8",
    )


if __name__ == "__main__":
    main()
